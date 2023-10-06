from openpyxl import *
from tkinter import *
from tkinter.ttk import Combobox

wb = load_workbook('ThongTinDangKyHocPhan.xlsx')

sheet = wb.active
def excel():
    sheet.column_dimensions['A'].width = 10
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 20
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 10
    sheet.column_dimensions['G'].width = 15
    sheet.column_dimensions['H'].width = 100
    sheet.cell(row=1, column=1).value = "Mã số sinh viên"
    sheet.cell(row=1, column=2).value = "Họ tên"
    sheet.cell(row=1, column=3).value = "Ngày sinh"
    sheet.cell(row=1, column=4).value = "Email"
    sheet.cell(row=1, column=5).value = "Số điện thoại"
    sheet.cell(row=1, column=6).value = "Học kỳ"
    sheet.cell(row=1, column=7).value = "Năm học"
    sheet.cell(row=1, column=8).value = "Môn học"

def focus1(event):
    hoTen_field.focus_set()
    
def focus2(event):
    ns_field.focus_set()

def focus3(event):
    email_field.focus_set()

def focus4(event):
    sdt_field.focus_set()
    
def focus5(event):
    hk_field.focus_set()

def focus6(event):
    namHoc_combo.focus_set()

def clear():
    name_field.delete(0, END)
    hoTen_field.delete(0, END)
    ns_field.delete(0, END)
    email_field.delete(0, END)
    sdt_field.delete(0, END)
    hk_field.delete(0, END)
    namHoc_combo.delete(0, END)

def insert():
    if (name_field.get() == "" and
        hoTen_field.get() == "" and
        ns_field.get() == "" and
        email_field.get() == "" and
        sdt_field.get() == "" and
        hk_field.get() == "" and
        namHoc_combo.get() == ""):
             
        print("empty input")
 
    else:
        current_row = sheet.max_row
        current_column = sheet.max_column
        selected_namhoc = namHoc_combo.get()
        sheet.cell(row=current_row + 1, column=1).value = name_field.get()
        sheet.cell(row=current_row + 1, column=2).value = hoTen_field.get()
        sheet.cell(row=current_row + 1, column=3).value = ns_field.get()
        sheet.cell(row=current_row + 1, column=4).value = email_field.get()
        sheet.cell(row=current_row + 1, column=5).value = sdt_field.get()
        sheet.cell(row=current_row + 1, column=6).value = hk_field.get()
        sheet.cell(row=current_row + 1, column=7).value = selected_namhoc

        subjects = []
        if python_var.get():
            subjects.append('Lập trình Python')
        if java_var.get():
            subjects.append('Lập trình Java')
        if cnpm_var.get():
            subjects.append('Công nghệ phần mềm')
        if web_var.get():
            subjects.append('Phát triển ứng dụng Web')
        sheet.cell(row=current_row + 1, column=8).value = ', '.join(subjects)
        
        wb.save('ThongTinDangKyHocPhan.xlsx')

        name_field.focus_set()
        
        clear()

if __name__ == "__main__":
    root = Tk()
    root.configure(background='light green')
    root.title("registration form")
    root.geometry("500x380") 
    excel()
    heading = Label(root, text="ĐĂNG KÝ HỌC PHẦN", bg="light green", font=("Arial", 14, "bold"))
    heading.grid(row=0, column=0, columnspan=4, padx=10, pady=10)

    labels = ["Mã số sinh viên", "Họ tên", "Ngày sinh", "Email", "Số điện thoại", "Học kỳ", "Năm học"]
    for i, label in enumerate(labels):
        lbl = Label(root, text=label, bg="light green")
        lbl.grid(row=i+1, column=0, sticky=W, padx=10, pady=5)
    
    monhoc_label = Label(root, text="Chọn môn học", bg="light green")
    monhoc_label.grid(row=8, column=0, sticky=W, padx=10, pady=5)

    python_var = IntVar()
    java_var = IntVar()
    cnpm_var = IntVar()
    web_var = IntVar()

    python_checkbox = Checkbutton(root, text="Lập trình Python", variable=python_var)
    java_checkbox = Checkbutton(root, text="Lập trình Java", variable=java_var)
    cnpm_checkbox = Checkbutton(root, text="Công nghệ phần mềm", variable=cnpm_var)
    ptudw_checkbox = Checkbutton(root, text="Phát triển ứng dụng Web", variable=web_var)

    python_checkbox.grid(row=8, column=1, sticky=W)
    java_checkbox.grid(row=8, column=1, sticky=E)
    cnpm_checkbox.grid(row=9, column=1, sticky=W)
    ptudw_checkbox.grid(row=9, column=1, sticky=E)

    entries = []
    for i in range(7):
        entry = Entry(root)
        entry.grid(row=i+1, column=1, ipadx="100")
        entries.append(entry)

    namHoc_values = ['2021-2022', '2022-2023', '2023-2024', '2024-2025', '2025-2026'] 
    namHoc_combo = Combobox(root, values=namHoc_values)
    namHoc_combo.grid(row=7, column=1, ipadx="90")
    entries.append(namHoc_combo)
    
    name_field, hoTen_field, ns_field, email_field, sdt_field, hk_field, namHoc_combo, *_ = entries
    
    name_field.bind("<Return>", focus1)
    hoTen_field.bind("<Return>", focus2)
    ns_field.bind("<Return>", focus3)
    email_field.bind("<Return>", focus4)
    sdt_field.bind("<Return>", focus5)
    hk_field.bind("<Return>", focus6)
    

    excel()
    submit = Button(root, text="Đăng ký", fg="black", bg="green", command=insert)
    submit.grid(row=10, column=0, pady=10)

    exit_button = Button(root, text="Thoát", fg="black", bg="green", command=root.quit)
    exit_button.grid(row=10, column=1, pady=10)
    root.mainloop()
